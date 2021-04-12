from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import time
import openpyxl
import sqlite3
import math
# 讀取資料庫
mydb = sqlite3.connect("TaxNo.db")
cursor = mydb.cursor()
cursor.execute("SELECT value FROM CrawlerInfos WHERE key='industryIndex'")
for row in cursor:
    industryIndex = int(row[0])

cursor.execute("SELECT value FROM CrawlerInfos WHERE key='companyPageIndex'")
for row in cursor:
    companyPageIndex = int(row[0])

cursor.execute(
    "SELECT value FROM CrawlerInfos WHERE key='companyPageInfoIndex'")
for row in cursor:
    companyPageInfoIndex = int(row[0])

cursor.execute("SELECT value FROM CrawlerInfos WHERE key='jobPageIndex'")
for row in cursor:
    jobPageIndex = int(row[0])

cursor.execute("SELECT value FROM CrawlerInfos WHERE key='jobInfoIndex'")
for row in cursor:
    jobInfoIndex = int(row[0])

options = Options()
chrome = webdriver.Chrome('./chromedriver', chrome_options=options)
chrome.get("http://www.104.com.tw/jb/104i/custlist/list?order=8")
chrome.implicitly_wait(30)
# 將產業列表展開
industryButton = chrome.find_element_by_xpath(
    "/html/body/div[3]/div[2]/div/div[1]/div[1]/h4")
industryButton.click()
for ind in range(industryIndex, 49, 1):
    # 獲取產業列表
    industryLinks = chrome.find_elements_by_class_name("a2")
    chrome.get(industryLinks[industryIndex].get_attribute('href'))
    # 取得總頁數
    resultCnt = int(chrome.find_element_by_class_name(
        'resultCnt').text)
    totalPage = math.ceil(resultCnt/15)
    # 取得當下網址
    currentUrl = str(chrome.current_url)
    # 跳至紀錄頁數
    recordUrl = currentUrl+"&page="+str(companyPageIndex)
    chrome.get(recordUrl)
    # 走訪產業下所有頁數
    for cc in range(companyPageIndex, totalPage, 1):
        companyPage = chrome.find_elements_by_class_name('page')
        if(len(companyPage) > 1 and cc > companyPageIndex):
            companyNextPageUrl = companyPage[1].get_attribute('href')
            chrome.get(companyNextPageUrl)
        elif(len(companyPage) > 0 and cc > companyPageIndex):
            if(companyPage[0].text == '下一頁 »'):
                companyNextPageUrl = companyPage[0].get_attribute('href')
                chrome.get(companyNextPageUrl)
        # 獲取單頁所有公司資連結
        companyInfos = chrome.find_elements_by_class_name("a5")
        allJobSoup = BeautifulSoup(
            chrome.page_source, 'html.parser')
        allJobList = allJobSoup.find_all(
            'a', {'class': 'a2', 'target': '_blank'})
        # 走訪單頁公司資訊__以下還需加入公司職缺頁數紀錄
        currentCompanyPageInfoIndex = 1
        for c in companyInfos:
            if(currentCompanyPageInfoIndex >= companyPageInfoIndex):
                # 開啟公司資訊分頁
                print(c.get_attribute('href'))
                js = "window.open('"+c.get_attribute('href')+"')"
                chrome.execute_script(js)
                # 切換窗口
                allHandles = chrome.window_handles
                routeWindowHandle = 0
                routeTitle = 'None'
                oringalHandles = allHandles[0]
                newHandles = allHandles[1]
                chrome.switch_to_window(newHandles)
                routeTitle = chrome.title
                # 取得公司資訊
                companySoup = BeautifulSoup(
                    chrome.page_source, 'html.parser')
                companyInfos = companySoup.find_all(
                    'span', {'class': 'condition'})
                companyName = chrome.find_element_by_xpath(
                    "/html/body/div[3]/div[1]/div/h1").text
                industry = companyInfos[0].text.strip()
                employeeCount = companyInfos[1].text.strip()
                captial = companyInfos[2].text.strip()
                contactPerson = companyInfos[3].text.strip()
                address = companyInfos[4].text.strip()
                companyUrl = companyInfos[7].text.strip()
                # 查詢統編
                cursor.execute(
                    "SELECT taxNo FROM Company WHERE companyName Like '%"+companyName+"%'")
                for row in cursor:
                    taxNo = str(row[0])
                # 連結至所有工作機會
                jobUrl = "http://www.104.com.tw"
                jobUrl += allJobList[currentCompanyPageInfoIndex]['href']
                chrome.get(jobUrl)
                # 撈取所有工作機會
                jobSoup = BeautifulSoup(
                    chrome.page_source, 'html.parser')
                # 取得頁數資訊page
                page = chrome.find_elements_by_class_name('page')
                currentPage = chrome.find_elements_by_class_name(
                    'fontStrength')
                if(len(page) > 0):
                    while(page[0].text == "下一頁 »"):
                        currentPage = currentPage[0].text
                        nextPageUrl = page[0].get_attribute('href')
                        chrome.get(nextPageUrl)
                        page = chrome.find_elements_by_class_name('page')
                        nextPageUrl = page[0].get_attribute('href')
                        chrome.get(nextPageUrl)
                        page = jobSoup.find_all(
                            'a', {'class': 'page'})

                        jobInfos = jobSoup.find_all(
                            'div', {'itemtype': 'http://schema.org/JobPosting'})
                        currentJobInfoIndex = 1
                        for j in jobInfos:
                            if(currentJobInfoIndex >= jobInfoIndex):
                                job = j.text.split()
                                jobTitle = job[0]
                                jobDate = job[len(job)-1]
                                # 寫入excel
                                fn = '104.xlsx'
                                wb = openpyxl.load_workbook(fn)
                                wb.active = 0
                                ws = wb.active
                                newRow = ws.max_row+1
                                ws.cell(
                                    column=1, row=newRow).value = companyName
                                ws.cell(column=2, row=newRow).value = taxNo
                                ws.cell(column=3, row=newRow).value = industry
                                ws.cell(
                                    column=4, row=newRow).value = contactPerson
                                ws.cell(column=5, row=newRow).value = captial
                                ws.cell(
                                    column=6, row=newRow).value = employeeCount
                                ws.cell(column=7, row=newRow).value = address
                                ws.cell(
                                    column=8, row=newRow).value = jobTitle+" "+jobDate
                                ws.cell(column=9, row=newRow).value = companyUrl
                                wb.save(fn)
                                jobInfoIndex += 1
                                # 更新jobInfoIndex
                                cursor.execute("UPDATE CralwerInfos set value =" +
                                               str(jobInfoIndex)+" where key='jobInfoIndex'")
                                mydb.commit()
                            currentJobInfoIndex += 1
                    # 重置jobInfoIndex
                    cursor.execute(
                        "UPDATE CralwerInfos set value ='1' where key='jobInfoIndex'")
                    mydb.commit()
                    jobInfoIndex = 1

                    jobPageIndex += 1
                    # 更新jobPageIndex
                    cursor.execute("UPDATE CralwerInfos set value =" +
                                   str(jobPageIndex)+" where key='jobPageIndex'")
                    mydb.commit()
                chrome.close()
                chrome.switch_to_window(oringalHandles)
            currentCompanyPageInfoIndex += 1
            companyPageInfoIndex += 1
            # 更新companyPageInfoIndex
            cursor.execute("UPDATE CralwerInfos set value =" +
                           str(companyPageInfoIndex)+" where key='companyPageInfoIndex'")
            mydb.commit()
        # 更新industryIndex
        cursor.execute("UPDATE CralwerInfos set value =" +
                       str(companyPageIndex+1)+" where key='companyPageIndex'")
        mydb.commit()

    # 更新industryIndex
    cursor.execute("UPDATE CralwerInfos set value =" +
                   str(industryIndex+1)+" where key='industryIndex'")
    mydb.commit()
