import sqlite3
import openpyxl

mydb = sqlite3.connect("Company.db")
cursor = mydb.cursor()
# 開啟工作簿
wb = openpyxl.load_workbook('GCIS_Company.xlsx')
# 獲取表單
sh = wb['工作表1']

# 獲取最大行數
print(sh.max_row)
# 獲取最大列數print(sh.max_column)

# 按行讀取所有資料，每一行的單元格放入一個元組中
rows = sh.rows
# 我們可以通過for迴圈以及value來檢視單元格的值
for row in list(rows):  # 遍歷每行資料
    companyName=str(row[0].value)
    companyTaxNo=str(row[1].value)
    companyStatus=str(row[2].value)
    companyCapital=str(row[3].value)
    companyPerson=str(row[4].value)
    companyAddress=str(row[5].value)
    paperYear=str(row[6].value)
    
    cursor.execute(
    "INSERT INTO CompanyData VALUES ('"+companyName+"','"+companyTaxNo+"','"+companyStatus+"','"+companyCapital+"','"+companyPerson+"','"+companyAddress+"','"+paperYear+"','');")                                   
    mydb.commit()