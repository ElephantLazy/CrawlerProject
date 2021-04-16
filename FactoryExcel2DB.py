import sqlite3
import openpyxl

mydb = sqlite3.connect("Factory.db")
cursor = mydb.cursor()
# 開啟工作簿
wb = openpyxl.load_workbook('GCIS_Factory.xlsx')
# 獲取表單
sh = wb['工作表1']

# 獲取最大行數
print(sh.max_row)
# 獲取最大列數print(sh.max_column)

# 按行讀取所有資料，每一行的單元格放入一個元組中
rows = sh.rows
# 我們可以通過for迴圈以及value來檢視單元格的值
for row in list(rows):  # 遍歷每行資料
    factoryName=str(row[0].value)
    factoryTaxNo=str(row[1].value)
    factoryCapital=str(row[2].value)
    factoryAddress=str(row[3].value)
    factoryStatus=str(row[4].value)
    paperYear=str(row[5].value)
    cursor.execute(
    "INSERT INTO FactoryData VALUES ('"+factoryName+"','"+factoryTaxNo+"','"+factoryCapital+"','"+factoryAddress+"','"+factoryStatus+"','"+paperYear+"','');")
    mydb.commit()