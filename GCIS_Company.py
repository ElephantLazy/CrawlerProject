from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.support.ui import Select
from bs4 import BeautifulSoup
import time
import openpyxl
import sqlite3
from retrying import retry


@retry
def crawlerCompany():
    # 讀取資料庫
    mydb = sqlite3.connect("Company.db")
    cursor = mydb.cursor()
    cursor.execute("SELECT value FROM GCIS_Company WHERE key='pageIndex'")
    for row in cursor:
        pageIndex = int(row[0])
    cursor.execute("SELECT value FROM GCIS_Company WHERE key='infoPageIndex'")
    for row in cursor:
        infoPageIndex = int(row[0])
    cursor.execute("SELECT value FROM GCIS_Company WHERE key='cityIndex'")
    for row in cursor:
        cityIndex = int(row[0])
    cursor.execute(
        "SELECT value FROM GCIS_Company WHERE key='mainSelectIndex'")
    for row in cursor:
        mainSelectIndex = int(row[0])
    cursor.execute("SELECT value FROM GCIS_Company WHERE key='selectIndex'")
    for row in cursor:
        selectIndex = int(row[0])

    options = Options()
    # options.add_argument("--disable-notifications")
    chrome = webdriver.Chrome('./chromedriver', chrome_options=options)
    chrome.get(
        "https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do;jsessionid=6AC3B69556A77E80B2D953B9C94F6B81")
    chrome.implicitly_wait(60)
    # 獲取資料種類選項
    checkBoxs = chrome.find_elements_by_name("qryType")
    checkBoxs[1].click()
    # 獲取登記現況選項
    statusBoxs = chrome.find_elements_by_name("isAlive")
    time.sleep(1)
    statusBoxs[0].click()
    js = "javascript:toggleAdvSearch('turnOn')"
    chrome.execute_script(js)
    # 填入搜尋欄
    searchBox = chrome.find_element_by_id('qryCond')
    searchBox.send_keys('公司')
    # 取得鎖營事業
    mainSelect = Select(chrome.find_element_by_id('busiItemMain'))
    mainSelectCount = len(mainSelect.options)
    isFirstCycle = True
    for cityIndex in range(cityIndex, 22, 1):
        cityBoxs = chrome.find_elements_by_name("city")
        cityBoxs[cityIndex].click()
        if(isFirstCycle == False):
            cityBoxs[cityIndex-1].click()
        isFirstCycle = False
        # 走訪所營事業
        for mainSelectIndex in range(mainSelectIndex, len(mainSelect.options), 1):
            mainSelect = Select(chrome.find_element_by_id('busiItemMain'))
            mainSelect.select_by_index(mainSelectIndex)
            select = Select(chrome.find_element_by_id('busiItemSub'))
            selectCount = len(select.options)
            # 走訪所營事業細項
            for selectIndex in range(selectIndex, len(select.options), 1):
                select = Select(chrome.find_element_by_id('busiItemSub'))
                select.select_by_index(selectIndex)
                time.sleep(1)
                # 送出搜尋
                js = "sendQueryList()"
                time.sleep(1)
                chrome.execute_script(js)
                # 取得分頁數量
                pageSoup = BeautifulSoup(chrome.page_source, 'html.parser')
                pageInfo = pageSoup.find_all('div', {
                    'style': 'float:left'})
                if(len(pageInfo) > 0):
                    startIndex = pageInfo[0].text.find('分')
                    endIndex = pageInfo[0].text.find('頁')
                    pageCount = int(pageInfo[0].text[startIndex+1:endIndex])
                    for i in range(pageIndex, pageCount+1, 1):
                        if(pageCount > 1):
                            js = "gotoPage("+str(i)+");"
                            time.sleep(2)
                            chrome.execute_script(js)
                        # 取得單頁所有公司詳細資訊
                        factorySoup = BeautifulSoup(
                            chrome.page_source, 'html.parser')
                        factoryInfos = factorySoup.find_all(
                            'span', {'class': 'moreLinkMouseOut'})
                        currentInfoIndex = 0
                        for f in factoryInfos:
                            if(currentInfoIndex >= infoPageIndex-1):
                                # 進入詳細資訊頁面
                                js = f.get("onclick")
                                time.sleep(2)
                                chrome.execute_script(js)
                                # 取得頁面資訊
                                companyDetailSoup = BeautifulSoup(
                                    chrome.page_source, 'html.parser')
                                companyDetail = companyDetailSoup.find_all(
                                    'td')
                                for c in range(0, 35, 1):
                                    if(companyDetail[c].text == '統一編號'):
                                        companyTaxNo = companyDetail[c+1].text.lstrip().rstrip()[
                                            0:8]
                                    if(companyDetail[c].text == '公司名稱'):
                                        endIndex = companyDetail[c+1].text.lstrip().find(
                                            '\n')
                                        companyName = companyDetail[c +
                                                                    1].text[0:endIndex]
                                    if(companyDetail[c].text == '公司所在地'):
                                        endIndex = companyDetail[c+1].text.lstrip().find(
                                            '\n')
                                        companyAddress = companyDetail[c+1].text.lstrip(
                                        ).rstrip()[0:endIndex]
                                    if(companyDetail[c].text == '代表人姓名'):
                                        companyPerson = companyDetail[c+1].text.lstrip(
                                        ).rstrip()
                                    if(companyDetail[c].text == '資本總額(元)'):
                                        companyCapital = companyDetail[c+1].text.lstrip(
                                        ).rstrip()
                                    if(companyDetail[c].text == '公司狀況'):
                                        endIndex = companyDetail[c+1].text.lstrip().find(
                                            '\n')
                                        companyStatus = companyDetail[c+1].text.lstrip(
                                        ).rstrip()[0:endIndex]
                                    if(companyDetail[c].text == '最後核准變更日期'):
                                        paperYear = companyDetail[c+1].text.lstrip(
                                        ).rstrip()

                                # 寫入excel
                                fn = 'GCIS_Company.xlsx'
                                wb = openpyxl.load_workbook(fn)
                                wb.active = 0
                                ws = wb.active
                                newRow = ws.max_row+1
                                ws.cell(
                                    column=1, row=newRow).value = companyName
                                ws.cell(
                                    column=2, row=newRow).value = companyTaxNo
                                ws.cell(
                                    column=3, row=newRow).value = companyStatus
                                ws.cell(
                                    column=4, row=newRow).value = companyCapital
                                ws.cell(
                                    column=5, row=newRow).value = companyPerson
                                ws.cell(
                                    column=6, row=newRow).value = companyAddress
                                ws.cell(column=7, row=newRow).value = paperYear
                                wb.save(fn)
                                # 返回上一頁
                                js = "javascript:history.back();"
                                chrome.execute_script(js)
                                infoPageIndex += 1
                                # 更新infoPageIndex
                                cursor.execute("UPDATE GCIS_Company set value =" +
                                               str(infoPageIndex)+" where key='infoPageIndex'")
                                mydb.commit()
                            currentInfoIndex += 1
                        # 更新infoPageIndex
                        infoPageIndex = 1
                        cursor.execute("UPDATE GCIS_Company set value =" +
                                       str(infoPageIndex)+" where key='infoPageIndex'")
                        mydb.commit()
                        cursor.execute("UPDATE GCIS_Company set value =" +
                                       str(pageIndex+1)+" where key='pageIndex'")
                        mydb.commit()
                        pageIndex += 1

                # 重置pageIndex
                pageIndex = 1
                cursor.execute(
                    "UPDATE GCIS_Company set value ='1' where key='pageIndex'")
                mydb.commit()
                cursor.execute("UPDATE GCIS_Company set value =" +
                               str(selectIndex+1)+" where key='selectIndex'")
                mydb.commit()
            # 重置selectIndex
            cursor.execute(
                "UPDATE GCIS_Company set value ='1' where key='selectIndex'")
            mydb.commit()
            selectIndex = 1
            cursor.execute("UPDATE GCIS_Company set value =" +
                           str(mainSelectIndex+1)+" where key='mainSelectIndex'")
            mydb.commit()

        # 重置mainSelectIndex
        cursor.execute(
            "UPDATE GCIS_Company set value ='1' where key='mainSelectIndex'")
        mydb.commit()
        mainSelectIndex = 1
        # 更新cityIndex
        cursor.execute("UPDATE GCIS_Company set value =" +
                       str(cityIndex+1)+" where key='cityIndex'")
        mydb.commit()


if __name__ == '__main__':
    crawlerCompany()
