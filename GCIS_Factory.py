from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from bs4 import BeautifulSoup
import time
import openpyxl
import sqlite3
from retrying import retry

@retry
def crawlerFactory():
    while True:
        try:
            # 讀取資料庫
            mydb = sqlite3.connect("Factory.db")
            cursor = mydb.cursor()
            cursor.execute("SELECT value FROM GCIS_Factory WHERE key='pageIndex'")
            for row in cursor:
                pageIndex = int(row[0])
            cursor.execute("SELECT value FROM GCIS_Factory WHERE key='infoPageIndex'")
            for row in cursor:
                infoPageIndex = int(row[0])
            cursor.execute("SELECT value FROM GCIS_Factory WHERE key='cityIndex'")
            for row in cursor:
                cityIndex = int(row[0])

            options = Options()
            # options.add_argument('--headless')  #規避google bug
            # options.add_argument('--disable-gpu')
            chrome = webdriver.Chrome('./chromedriver', chrome_options=options)
            chrome.get("https://findbiz.nat.gov.tw/fts/query/QueryBar/queryInit.do;jsessionid=6AC3B69556A77E80B2D953B9C94F6B81")
            chrome.implicitly_wait(60)
            # 獲取資料種類選項
            checkBoxs = chrome.find_elements_by_name("qryType")
            checkBoxs[3].click()
            time.sleep(1)
            checkBoxs[0].click()
            # 獲取登記現況選項
            statusBoxs = chrome.find_elements_by_name("isAlive")
            time.sleep(1)
            statusBoxs[0].click()
            js = "javascript:toggleAdvSearch('turnOn')"
            chrome.execute_script(js)
            # 填入搜尋欄
            searchBox = chrome.find_element_by_id('qryCond')
            searchBox.send_keys('~廠')
            # 獲取城市選項
            isFirstCycle = True
            for cityIndex in range(cityIndex, 22, 1):
                cursor.execute("UPDATE GCIS_Factory set value =" +
                            str(cityIndex)+" where key='cityIndex'")
                mydb.commit()
                time.sleep(1)
                cityBoxs = chrome.find_elements_by_name("city")
                cityBoxs[cityIndex].click()
                if(isFirstCycle == False):
                    cityBoxs[cityIndex-1].click()
                isFirstCycle = False
                # 送出搜尋
                js = "sendQueryList()"
                time.sleep(2)
                chrome.execute_script(js)
                # 取得分頁數量
                pageSoup = BeautifulSoup(chrome.page_source, 'html.parser')
                pageInfo = pageSoup.find_all('div', {
                    'style': 'float:left'})
                startIndex = pageInfo[0].text.find('分')
                endIndex = pageInfo[0].text.find('頁')
                pageCount = int(pageInfo[0].text[startIndex+1:endIndex])
                for i in range(pageIndex, pageCount, 1):
                    js = "gotoPage("+str(i)+");"
                    time.sleep(2)
                    chrome.execute_script(js)
                    # 取得單頁所有工廠詳細資訊
                    factorySoup = BeautifulSoup(chrome.page_source, 'html.parser')
                    factoryInfos = factorySoup.find_all(
                        'span', {'class': 'moreLinkMouseOut'})
                    currentInfoIndex = 0
                    for f in factoryInfos:
                        if(currentInfoIndex >= infoPageIndex-1):
                            # 進入詳細資訊頁面
                            js = f.get("onclick")
                            time.sleep(2)
                            chrome.execute_script(js)
                            # 取得頁面資訊
                            factoryDetailSoup = BeautifulSoup(
                                chrome.page_source, 'html.parser')
                            factoryDetail = factoryDetailSoup.find_all('td')
                            endIndex = factoryDetail[3].text.find('\n')
                            factoryName = factoryDetail[3].text[0:endIndex]
                            factoryAddress = factoryDetail[13].contents[0].lstrip(
                            ).rstrip()
                            factoryCapital = factoryDetail[69].text.lstrip().rstrip()
                            factoryTaxNo = factoryDetail[17].text.lstrip().rstrip()[0:8]
                            factoryStatus = factoryDetail[23].text.lstrip().rstrip()
                            paperYear = factoryDetail[83].text.lstrip().rstrip()
                            # 寫入excel
                            # fn = 'GCIS_Factory.xlsx'
                            # wb = openpyxl.load_workbook(fn)
                            # wb.active = 0
                            # ws = wb.active
                            # newRow = ws.max_row+1
                            # ws.cell(column=1, row=newRow).value = factoryName
                            # ws.cell(column=2, row=newRow).value = factoryTaxNo
                            # ws.cell(column=3, row=newRow).value = factoryCapital
                            # ws.cell(column=4, row=newRow).value = factoryAddress
                            # ws.cell(column=5, row=newRow).value = factoryStatus
                            # ws.cell(column=6, row=newRow).value = paperYear
                            # wb.save(fn)

                            #Write to DB
                            cursor.execute(
                                "INSERT INTO FactoryData VALUES ('"+factoryName+"','"+factoryTaxNo+"','"+factoryCapital+"','"+factoryAddress+"','"+factoryStatus+"','"+paperYear+"','');")
                            mydb.commit()
                            # 返回上一頁
                            js = "javascript:history.back();"
                            chrome.execute_script(js)
                            infoPageIndex += 1
                            # 更新infoPageIndex
                            cursor.execute("UPDATE GCIS_Factory set value =" +
                                        str(infoPageIndex)+" where key='infoPageIndex'")
                            mydb.commit()
                        currentInfoIndex += 1
                    # 更新infoPageIndex
                    infoPageIndex = 1
                    cursor.execute("UPDATE GCIS_Factory set value =" +
                                str(infoPageIndex)+" where key='infoPageIndex'")
                    mydb.commit()
                    cursor.execute("UPDATE GCIS_Factory set value =" +
                                str(pageIndex+1)+" where key='pageIndex'")
                    mydb.commit()
                    pageIndex += 1

                # 重置pageIndex
                pageIndex = 1
                cursor.execute("UPDATE GCIS_Factory set value ='1' where key='pageIndex'")
                mydb.commit()
        finally:
            if('chrome' in globals()):
                chrome.quit()

if __name__ == '__main__':
    crawlerFactory()
