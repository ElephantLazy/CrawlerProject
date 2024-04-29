import sqlite3
import openpyxl

# 開啟目標資料庫，這裡以db1為例
mydb = sqlite3.connect("./var/db1.db")
cursor = mydb.cursor()

# 開啟工作簿
wb = openpyxl.load_workbook('D:\CrawlerProject-main\src\data.xlsx')

# 獲取表單
sh = wb['BGMOPEN1']

# 獲取最大行數
print(sh.max_row)

# 獲取最大列數
print(sh.max_column)

# 按行讀取所有資料，每一行的單元格放入一個元組中
rows = sh.rows

# 資料庫計數器
db_counter = 1
# 資料庫最大筆數
max_records_per_db = 7468
# 目前處理的筆數
record_count = 0


# 我們可以通過for迴圈以及value來檢視單元格的值
for row in list(rows):
    sql_template = "INSERT INTO CrawlerData VALUES ('{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}','{}');"

    # 將row中的每個值插入到模板中，如果值為None，則轉換為空白字串
    sql_query = sql_template.format(
        row[0].value if row[0].value is not None else '',
        row[1].value if row[1].value is not None else '',
        row[2].value if row[2].value is not None else '',
        row[3].value if row[3].value is not None else '',
        row[4].value if row[4].value is not None else '',
        row[5].value if row[5].value is not None else '',
        row[6].value if row[6].value is not None else '',
        row[7].value if row[7].value is not None else '',
        '',
        '',
        '',
        '',
        '',
        '',
        '',
        ''
    )
    
    cursor.execute(sql_query)
    record_count += 1

    # 如果達到一萬筆資料，就切換到下一個資料庫
    if record_count == max_records_per_db:
        mydb.commit()
        mydb.close()
        db_counter += 1
        if db_counter <= 15:
            mydb = sqlite3.connect(f"./var/db{db_counter}.db")
            cursor = mydb.cursor()
            record_count = 0

# 確保最後一批資料被儲存
mydb.commit()
mydb.close()
