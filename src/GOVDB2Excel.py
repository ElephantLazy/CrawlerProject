import sqlite3
import openpyxl

mydb = sqlite3.connect("./var/GOV.db")
cursor = mydb.cursor()
cursor.execute("SELECT * FROM CompanyData")
for row in cursor:
    # 寫入excel
    fn = './src/GOV.xlsx'
    wb = openpyxl.load_workbook(fn)
    wb.active = 0
    ws = wb.active
    newRow = ws.max_row+1
    ws.cell(column=1, row=newRow).value = row[0]
    ws.cell(column=2, row=newRow).value = row[1]
    ws.cell(column=3, row=newRow).value = row[2]
    ws.cell(column=4, row=newRow).value = row[3]
    ws.cell(column=5, row=newRow).value = row[4]
    ws.cell(column=6, row=newRow).value = row[5]
    ws.cell(column=7, row=newRow).value = row[6]
    ws.cell(column=8, row=newRow).value = row[7]
    ws.cell(column=9, row=newRow).value = row[8]
    ws.cell(column=10, row=newRow).value = row[9].replace("\r", "").replace("\n", "").replace("\t", "")
    wb.save(fn)

